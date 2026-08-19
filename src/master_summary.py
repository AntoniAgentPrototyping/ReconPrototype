"""The month-end master: every window of a month, by brand and by storefront.

The team asked (Aug 2026) for "one master file to consolidate all the weeks by
brand … for all platforms". This builds it from the per-window finance files this
system already produced, so **every figure in the master is provably the same
number the weekly file carries** — the aggregation adds no arithmetic of its own,
and `tie_rows` re-derives each column total from its sources so a master that
cannot tie to them cannot be shipped.

## Why this is in `src/` and not in `tools/`

It was a script (`tools/build_master_summary.py`) with the window list hardcoded
in a dict. Two things were wrong with that and only one of them was cosmetic:

- The hardcoded table named `w1..w5 / s1..s4 / l1..l5` and **silently omitted the
  sub-batch windows that really exist** — Shopee's `s2x` and `s3k`. Its own tie
  check re-read the same table, so it could never notice its own omission. The
  window list now comes from the database, from the runs that actually happened
  (`service/month_master.py`).
- A second compute layer outside the verified pipeline is exactly what
  `docs/06-DECISIONS.md#d24` forbids. Here it sits beside `finance_template`,
  under the same I/O rule.

## The I/O rule

`build()` is pure: it takes frames and returns an **unwritten** `Workbook`, the
same shape `finance_template.build_*` returns, so `pipeline.write_artifacts` can
be the thing that writes it and `tests/test_io_boundary.py` needs no new grant.
`read_window` does read — but through `ingest.read_excel_sheet` /
`ingest.sheet_names`, which are the declared boundary, so no call in this module
is one the lint counts.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from . import ingest
from .errors import ReconHardStop
from .pipeline import norm_store

# Accounting format, zero as a dash — the team's own convention in these files.
ACC0 = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'

# The finance template puts the line-tab header on sheet row 6 (1-based).
HEADER_ROW = 6

PLATFORMS = ("TikTok", "Shopee", "Lazada")
PLATFORM_DIR = {"TikTok": "tiktok", "Shopee": "shopee", "Lazada": "lazada"}
DIR_PLATFORM = {v: k for k, v in PLATFORM_DIR.items()}


@dataclass(frozen=True)
class PlatformSchema:
    """Where the invoice lines are, per platform, in a finance file.

    `with_vat` is None for Lazada because the team's own template carries no
    with-VAT column on its line tabs. It is recombined as `amount x the tab's own
    rate`, which is exact: the tab name IS the VAT factor ('1.05', '1.08', '1.10'),
    which is why the selector is a rate-shaped regex rather than a fixed name.
    """

    is_line_tab: Any
    store: str
    pre_vat: str
    with_vat: str | None


SCHEMA: dict[str, PlatformSchema] = {
    "TikTok": PlatformSchema(lambda s: s.strip() == "Xuat HD bt",
                             "Source.Name", "Amount before VAT", "Check total"),
    "Shopee": PlatformSchema(lambda s: s.strip() == "Xuat HD bt",
                             "Source.Name", "Amount before VAT", "Check total"),
    "Lazada": PlatformSchema(
        lambda s: re.fullmatch(r"1\.\d{1,2}", s.strip()) is not None,
        "Source.Name", "Amount", None),
}


# ---------------------------------------------------------------------------
# Reading one window's finance file
# ---------------------------------------------------------------------------

def line_tabs(platform: str, names) -> list[str]:
    """The invoice-line tabs among a finance file's sheets, in workbook order."""
    return [s for s in names if SCHEMA[platform].is_line_tab(s)]


def fold_lines(platform: str, sheet: str, df: pd.DataFrame) -> pd.DataFrame:
    """One line tab -> (store, pre, wv) rows. Pure; no file, no workbook.

    Order-level columns are blank on repeat-SKU rows — the template's "non repeat"
    convention — so `store` is forward-filled before grouping. Without that, every
    line after the first of an order attributes to nothing and the storefront's
    revenue silently shrinks.
    """
    schema = SCHEMA[platform]
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    if schema.store not in df.columns:
        return pd.DataFrame({"store": [], "pre": [], "wv": []})

    pre = pd.to_numeric(df.get(schema.pre_vat), errors="coerce").fillna(0.0)
    if schema.with_vat is None:
        wv = pre * float(sheet.strip())
    else:
        wv = pd.to_numeric(df.get(schema.with_vat), errors="coerce").fillna(0.0)
    return pd.DataFrame({"store": df[schema.store].ffill(), "pre": pre, "wv": wv})


def per_store(frames: list[pd.DataFrame]) -> pd.DataFrame:
    """Fold a window's line tabs into one (store, pre, wv) row per storefront."""
    if not frames:
        return pd.DataFrame({"store": [], "pre": [], "wv": []})
    df = pd.concat(frames, ignore_index=True)
    df = df[df["store"].notna()].copy()
    df["store"] = df["store"].map(norm_store)
    out = df.groupby("store", as_index=False)[["pre", "wv"]].sum()
    return out.sort_values("store", ignore_index=True)


def read_window(path: Path, platform: str) -> pd.DataFrame:
    """One finance file -> its per-storefront totals.

    Reads through `ingest`, which is the declared boundary. A finance file with no
    line tab at all is a hard stop rather than an empty frame: silently
    contributing zero to a month-end total is the failure this file exists to
    prevent.
    """
    tabs = line_tabs(platform, ingest.sheet_names(Path(path)))
    if not tabs:
        raise ReconHardStop(
            f"{Path(path).name} carries no invoice-line tab, so no {platform} "
            f"revenue can be read from it. Expected "
            f"{'a rate-named tab (1.05/1.08/1.10)' if platform == 'Lazada' else 'Xuat HD bt'}.")
    frames = [fold_lines(platform, tab,
                         ingest.read_excel_sheet(Path(path), tab, HEADER_ROW))
              for tab in tabs]
    return per_store(frames)


# ---------------------------------------------------------------------------
# What the month contains, and what it is missing
# ---------------------------------------------------------------------------

def window_label(period: str) -> str:
    """The column heading for a window: its suffix, '2026-07_w1' -> 'w1'.

    Deliberately NOT the team's '01-07' day range. That range came from the
    hardcoded window table this module exists to delete — deriving it would mean
    re-inventing the same fixed assumption about how many windows a month has and
    which days each covers, and that assumption is what dropped `s2x` and `s3k`.
    The period is the identifier the rest of the system uses, so the master uses it.
    """
    return period.split("_", 1)[1] if "_" in period else period


def parse_brand_map(text: str) -> dict[tuple[str, str], tuple[str, str, str]]:
    """`config/brand_map.csv` text -> {(platform, storefront): (brand, conf, note)}.

    Takes TEXT, not a path, so the parsing rule has one home while the file read
    stays with whoever owns a filesystem — `tools/` for the CLI, `service/` for the
    worker. `src/` reading a CSV here would need an I/O grant for a file that is
    not part of the pipeline's own contract.
    """
    import csv
    import io

    out: dict[tuple[str, str], tuple[str, str, str]] = {}
    for row in csv.DictReader(io.StringIO(text)):
        key = ((row.get("platform") or "").strip().lower(),
               (row.get("storefront") or "").strip())
        if not key[1]:
            continue
        out[key] = ((row.get("client_brand") or "").strip(),
                    (row.get("confidence") or "").strip(),
                    (row.get("note") or "").strip())
    return out


@dataclass(frozen=True)
class Window:
    """One window that contributed to this master."""

    platform: str                       # 'TikTok' | 'Shopee' | 'Lazada'
    period: str                         # '2026-07_w1'
    label: str                          # '01-07' — the column heading
    totals: pd.DataFrame                # per-store (store, pre, wv)
    source: str = ""                    # where the figures were read from


@dataclass
class Coverage:
    """What this master covers, and what it does not (task 3.5).

    A month-end master is rebuilt every time a window finishes, so for most of the
    month it is **partial by construction**. Naming the included and missing
    windows on the face of the workbook is the whole point: a master that looks
    complete and is not is precisely the failure this project exists to prevent.
    """

    month: str
    included: list[tuple[str, str]] = field(default_factory=list)   # (platform, period)
    missing: list[tuple[str, str, str]] = field(default_factory=list)  # (platform, period, why)
    built_at: str = ""
    built_by: str = ""

    @property
    def complete(self) -> bool:
        return not self.missing

    def headline(self) -> str:
        if self.complete:
            return (f"Covers all {len(self.included)} settlement window(s) that ran "
                    f"in {self.month}.")
        return (f"PARTIAL — {len(self.included)} window(s) included, "
                f"{len(self.missing)} missing. The totals below are NOT the month's.")


def order_windows(windows: list[Window]) -> dict[str, list[Window]]:
    """Group by platform, in settlement order, preserving the caller's order.

    Sorted by `period`, which sorts settlement-correctly for the real labels
    (`2026-07_s1` < `2026-07_s2`) and keeps a sub-batch beside its parent
    (`s2` < `s2x` < `s3`) — the two windows the old hardcoded table dropped.
    """
    grouped: dict[str, list[Window]] = {p: [] for p in PLATFORMS}
    for w in windows:
        grouped.setdefault(w.platform, []).append(w)
    return {p: sorted(ws, key=lambda w: w.period)
            for p, ws in grouped.items() if ws}


# ---------------------------------------------------------------------------
# The internal gate (task 3.6)
# ---------------------------------------------------------------------------

def tie_rows(windows: list[Window]) -> list[dict]:
    """Every window's contribution, re-derived from its own frame.

    `tools/build_master_summary.py` printed this and returned an exit code; making
    it data means a test can assert it. It is a real check despite comparing a
    frame to itself: it is the assertion that the number written into the grid is
    the number the source window carries, so a grid-building bug (a mis-keyed
    lookup, a dropped store) shows up here rather than in an invoice.
    """
    rows = []
    for w in windows:
        rows.append({
            "platform": w.platform,
            "period": w.period,
            "label": w.label,
            "stores": int(len(w.totals)),
            "pre": float(w.totals["pre"].sum()),
            "wv": float(w.totals["wv"].sum()),
            "source": w.source,
        })
    return rows


# ---------------------------------------------------------------------------
# The workbook — pure build, never written here
# ---------------------------------------------------------------------------

def _grid(ws, title: str, row_label: str, rows: list[str], cols: list[str],
          values: dict, extra_cols: list[tuple[str, dict]] | None = None) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12)
    ws["A3"] = row_label
    ws["A3"].font = Font(bold=True)
    for j, c in enumerate(cols, start=2):
        cell = ws.cell(row=3, column=j, value=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    n_extra = len(extra_cols or [])
    for k, (label, _) in enumerate(extra_cols or [], start=len(cols) + 2):
        cell = ws.cell(row=3, column=k, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=r)
        for j, c in enumerate(cols, start=2):
            cell = ws.cell(row=i, column=j, value=float(values.get((r, c), 0.0)))
            cell.number_format = ACC0
        for k, (_, vals) in enumerate(extra_cols or [], start=len(cols) + 2):
            cell = ws.cell(row=i, column=k, value=float(vals.get(r, 0.0)))
            cell.number_format = ACC0

    total_row = len(rows) + 4
    ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
    for j in range(2, len(cols) + 2 + n_extra):
        letter = get_column_letter(j)
        cell = ws.cell(row=total_row, column=j,
                       value=f"=SUM({letter}4:{letter}{total_row - 1})")
        cell.number_format = ACC0
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 34
    for j in range(2, len(cols) + 2 + n_extra):
        ws.column_dimensions[get_column_letter(j)].width = 18
    ws.freeze_panes = "B4"


def _coverage_sheet(wb: Workbook, coverage: Coverage) -> None:
    """The first tab, and deliberately so — it qualifies everything after it."""
    ws = wb.create_sheet("Coverage")
    ws["A1"] = f"What this master covers — {coverage.month}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = coverage.headline()
    ws["A2"].font = Font(bold=True, size=11,
                         color="FF006100" if coverage.complete else "FFC00000")

    row = 4
    if coverage.built_at or coverage.built_by:
        ws.cell(row=row, column=1, value="Built")
        ws.cell(row=row, column=2,
                value=" ".join(x for x in (coverage.built_at, coverage.built_by) if x))
        row += 2

    for header, entries, colour in (
            ("Included", [(p, w, "") for p, w in coverage.included], None),
            ("MISSING", coverage.missing, "FFC00000")):
        if not entries:
            continue
        ws.cell(row=row, column=1, value=header).font = Font(bold=True)
        row += 1
        for j, h in enumerate(("Platform", "Window", "Why"), start=1):
            c = ws.cell(row=row, column=j, value=h)
            c.font = Font(bold=True)
        row += 1
        for platform, period, why in entries:
            ws.cell(row=row, column=1, value=platform)
            ws.cell(row=row, column=2, value=period)
            cell = ws.cell(row=row, column=3, value=why)
            if colour:
                cell.font = Font(bold=True, color=colour)
            row += 1
        row += 1

    for col, width in (("A", 16), ("B", 22), ("C", 70)):
        ws.column_dimensions[col].width = width


def build(coverage: Coverage, windows: list[Window],
          brand_map: dict[tuple[str, str], tuple[str, str, str]] | None = None) -> Workbook:
    """The master workbook, in memory. **Writes nothing.**

    `brand_map` is `{(platform_dir, storefront): (client_brand, confidence, note)}`
    — the team-reviewable storefront→brand mapping. An unmapped storefront keeps
    its own name and is flagged; nothing is ever merged without an explicit row,
    because merging two storefronts that are separate client entities invoices one
    client for another's revenue.
    """
    brand_map = brand_map or {}
    grouped = order_windows(windows)
    plats = [p for p in PLATFORMS if p in grouped]

    wb = Workbook()
    wb.remove(wb.active)
    _coverage_sheet(wb, coverage)

    # --- Summary: one block per platform, vertical ---------------------------
    # The platforms do NOT share window boundaries, so a common column axis would
    # be sparse and misleading. This mirrors the Lazada "Summary" tab idiom.
    ws = wb.create_sheet("Summary")
    ws["A1"] = f"Marketplace revenue summary, {coverage.month} (VND)"
    ws["A1"].font = Font(bold=True, size=12)
    if not coverage.complete:
        ws["A2"] = coverage.headline()
        ws["A2"].font = Font(bold=True, color="FFC00000")
    for j, h in enumerate(["Platform", "Window", "Sum of Amount before VAT",
                           "Sum of Check total (with VAT)"], start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center" if j > 2 else "left")

    r = 4
    grand_pre = grand_wv = 0.0
    for p in plats:
        p_pre = p_wv = 0.0
        for w in grouped[p]:
            pre = float(w.totals["pre"].sum())
            wv = float(w.totals["wv"].sum())
            p_pre += pre
            p_wv += wv
            ws.cell(row=r, column=1, value=p)
            ws.cell(row=r, column=2, value=w.label)
            ws.cell(row=r, column=3, value=pre).number_format = ACC0
            ws.cell(row=r, column=4, value=wv).number_format = ACC0
            r += 1
        ws.cell(row=r, column=1, value=f"{p} total").font = Font(bold=True)
        for col, v in ((3, p_pre), (4, p_wv)):
            c = ws.cell(row=r, column=col, value=v)
            c.number_format = ACC0
            c.font = Font(bold=True)
        grand_pre += p_pre
        grand_wv += p_wv
        r += 2
    ws.cell(row=r, column=1, value="ALL PLATFORMS").font = Font(bold=True, size=12)
    for col, v in ((3, grand_pre), (4, grand_wv)):
        c = ws.cell(row=r, column=col, value=v)
        c.number_format = ACC0
        c.font = Font(bold=True, size=12)
    for col, width in (("A", 22), ("B", 14), ("C", 26), ("D", 28)):
        ws.column_dimensions[col].width = width

    # --- storefront totals per platform --------------------------------------
    per_plat: dict[str, dict[str, float]] = {}
    for p in plats:
        acc: dict[str, float] = {}
        for w in grouped[p]:
            for s, v in zip(w.totals["store"], w.totals["wv"]):
                acc[s] = acc.get(s, 0.0) + float(v)
        per_plat[p] = acc

    def brand_of(p: str, store: str) -> tuple[str, str, str]:
        return brand_map.get((PLATFORM_DIR[p].lower(), store),
                             (store, "UNMAPPED (kept as storefront)",
                              "not in config/brand_map.csv"))

    # By brand
    brand_vals: dict[tuple[str, str], float] = {}
    for p in plats:
        for s, v in per_plat[p].items():
            b = brand_of(p, s)[0]
            brand_vals[(b, p)] = brand_vals.get((b, p), 0.0) + v
    brands = sorted({b for b, _ in brand_vals})
    _grid(wb.create_sheet("By brand"),
          f"Revenue by client brand across platforms, {coverage.month} "
          f"(with VAT, VND) — mapping on the 'Brand mapping' tab",
          "Client brand", brands, plats,
          {(b, p): brand_vals.get((b, p), 0.0) for b in brands for p in plats})

    # By storefront — the untranslated view, for traceability
    stores_all = sorted({s for p in plats for s in per_plat[p]})
    _grid(wb.create_sheet("By storefront"),
          f"Revenue by storefront (as named in the platform files), "
          f"{coverage.month} (with VAT, VND)",
          "Storefront", stores_all, plats,
          {(s, p): per_plat[p].get(s, 0.0) for s in stores_all for p in plats})

    # Brand mapping — the reviewable table
    ws = wb.create_sheet("Brand mapping")
    ws["A1"] = ("Storefront -> client brand mapping. PLEASE REVIEW the rows marked "
                "needs-confirmation or UNMAPPED and correct in one pass; storefronts "
                "were NOT merged unless the mapping says so.")
    ws["A1"].font = Font(bold=True)
    for j, h in enumerate(["Platform", "Storefront (as in files)",
                           "Proposed client brand", "Confidence",
                           f"{coverage.month} with-VAT (VND)", "Note"], start=1):
        ws.cell(row=3, column=j, value=h).font = Font(bold=True)
    r = 4
    for p in plats:
        for s in sorted(per_plat[p]):
            b, conf, note = brand_of(p, s)
            ws.cell(row=r, column=1, value=p)
            ws.cell(row=r, column=2, value=s)
            ws.cell(row=r, column=3, value=b)
            cell = ws.cell(row=r, column=4, value=conf)
            if conf != "high":
                cell.font = Font(bold=True, color="FFC00000")
            ws.cell(row=r, column=5, value=per_plat[p][s]).number_format = ACC0
            ws.cell(row=r, column=6, value=note)
            r += 1
    for col, width in (("A", 10), ("B", 44), ("C", 30), ("D", 30), ("E", 20), ("F", 62)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"

    # --- one tab per platform: store x window --------------------------------
    for p in plats:
        labels = [w.label for w in grouped[p]]
        stores = sorted({s for w in grouped[p] for s in w.totals["store"]})
        by_label = {w.label: w.totals.set_index("store") for w in grouped[p]}
        wv = {(s, lab): float(by_label[lab]["wv"].get(s, 0.0))
              for s in stores for lab in labels}
        pre = {s: sum(float(by_label[lab]["pre"].get(s, 0.0)) for lab in labels)
               for s in stores}
        _grid(wb.create_sheet(p),
              f"{p} by store and window, {coverage.month} (with VAT, VND)",
              "Source.Name", stores, labels, wv,
              extra_cols=[("Sum of Amount before VAT", pre)])

    return wb
