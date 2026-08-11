"""Finance workbooks in the team's invoicing-template shape — cleaned up.

Layout evidence (cell coordinates refer to the team's May files):
- TikTok : "Tiktok result 01 to 17T5 For KA.xlsx"
- Shopee : "shopee result Sample For KA 01 to 10T05.xlsx"
- Lazada : "Laz result KA used 26_04T5 to 10T5.xlsx"

Principle: same tabs, same column names/order, same header-row position and
total-cell positions as the template, so the team can read it without
retraining — but every check value is COMPUTED by the verified engine and
written as a static number.

Two families of numbers, deliberately kept distinct (the template mixes
them and several of its checks silently broke):

1. LINE tabs carry the engine's exact amounts. Their control blocks compare
   exact line totals to an exact per-VAT-rate recombination — a drift of
   even a few VND means a row was dropped or edited. Tolerance 2,000
   ('Xuat HD bt'!P5 / R5).
2. PIVOT tabs (brand / VAT / "KA used") are the client-facing invoice view:
   per-SKU rounded average price x quantity, so every displayed line
   satisfies qty x unit = amount. Recombining rounded prices drifts from
   the exact books by design; the PV-sum / Summary checks measure exactly
   that drift, at the tolerance the team's WORKING verdicts use
   (TikTok 'PV sum'!G8 12,000 · Shopee 'PV sum'!J11 10,000 ·
   Lazada Summary!F17 2,000).

Template defects fixed, not copied:
- TikTok 'Xuat HD bt' control: sums a #REF! (L3) and multiplies the KAO
  bucket by 1.10 (N4=M4*1.1). Rebuilt as exact per-VAT-rate rows.
- Shopee 'PV sum'!G3 verdict reads blank E2 (always "OK"); its stated 2,000
  tolerance never actually applied. Wired to the real diff at 10,000 — the
  team's functioning side-block tolerance for the same comparison.
- Shopee income control O3 reads ' 1.08 KA'!F2, which points at an empty
  column whenever the pivot is re-arranged. Values are written directly.
- 'PV xuat HD' checks compared mismatched quantities (always failing,
  ignored). Rebuilt as exact-vs-exact at 1,000.
- Lazada '1.05'/'1.10' tabs carried stale January stamps ("Laz T12",
  "8T1 to 14T1"). Stamps come from the current window.
- Lazada 'X KA used'!J1 used tolerance 1,000 on a pivot-rounding check
  whose natural drift is ~2,000 (their own May file fails it and is booked
  anyway); Summary uses 2,000 for the same comparison. Both now use 2,000.

Verdict wording is the team's own, diacritics included.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell

from .runlog import RunLog

# Accounting masks copied from the team's cells (probe: 'Xuat HD bt' row 7).
ACC0 = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'
ACC2 = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'
PLAIN0 = "#,##0_);(#,##0)"
PLAIN2 = "#,##0.00_);(#,##0.00)"

VERDICT_OK = "ok có thể xuất HD"
VERDICT_BAD = "Cần check lại số có vấn đề"
PVSUM_OK = "OK"
PVSUM_BAD = "check lai sai roi"
RETURN_FULL = "Return full ko xuat HD"
RETURN_PARTIAL = "Return 1 phan phai xuat HD"

VAT_RATES = (1.05, 1.08, 1.10)

TOL_PIVOT_DRIFT_SHOPEE = 10000.0   # 'PV sum'!J11 — the working verdict
TOL_PIVOT_DRIFT_LAZADA = 2000.0    # Summary!F17
TOL_SKU_PIVOT = 1000.0             # 'PV xuat HD'!I1 intent
TOL_LAZ_LINE = 1000.0              # '1.08'!I2


def _verdict(diff: float, tol: float, ok: str = VERDICT_OK, bad: str = VERDICT_BAD) -> str:
    return ok if abs(diff) < tol else bad


def _bucket(store: str, buckets: list[tuple[str, str]], default: str) -> str:
    s = str(store).lower()
    for needle, name in buckets:
        if needle in s:
            return name
    return default


TIKTOK_BUCKETS = [("kao", "KAO 8"), ("merries", "Merries 8")]
SHOPEE_BUCKETS = [("curel", "Curel"), ("kao", "KAO"), ("merries", "Merries"), ("kate", "Kate")]
LAZADA_BUCKETS = [("curel", "Curel.xlsx"), ("kao", "KAO.xlsx"), ("merries", "Merries.xlsx")]


class _Tab:
    """Sparse control rows + an optional streamed data region, emitted in
    row order to a write-only worksheet."""

    def __init__(self, wb: Workbook, name: str):
        self.ws = wb.create_sheet(name)
        self.cells: dict[int, dict[int, tuple]] = {}

    def put(self, ref: str, value, fmt: str | None = None) -> None:
        col = "".join(ch for ch in ref if ch.isalpha())
        row = int("".join(ch for ch in ref if ch.isdigit()))
        cidx = 0
        for ch in col:
            cidx = cidx * 26 + (ord(ch.upper()) - 64)
        self.cells.setdefault(row, {})[cidx] = (value, fmt)

    def label_row(self, row: int, start_col: int, labels: list[str]) -> None:
        for i, lab in enumerate(labels):
            self.cells.setdefault(row, {})[start_col + i] = (lab, None)

    def _make_row(self, rowcells: dict[int, tuple]) -> list:
        out = []
        for c in range(1, max(rowcells) + 1):
            if c in rowcells:
                v, fmt = rowcells[c]
                wc = WriteOnlyCell(self.ws, value=v)
                if fmt:
                    wc.number_format = fmt
                out.append(wc)
            else:
                out.append(None)
        return out

    def emit(self, data: pd.DataFrame | None = None, data_start_row: int | None = None,
             fmts: list[str | None] | None = None, widths: dict[str, float] | None = None,
             freeze: str | None = None) -> None:
        for ref, w in (widths or {}).items():
            self.ws.column_dimensions[ref].width = w
        if freeze:
            self.ws.freeze_panes = freeze
        last_control = max(self.cells) if self.cells else 0
        top = data_start_row - 1 if data_start_row else last_control
        for r in range(1, top + 1):
            rowcells = self.cells.get(r)
            self.ws.append(self._make_row(rowcells) if rowcells else [])
        if data is None:
            return
        fmts = fmts or [None] * data.shape[1]
        for tup in data.itertuples(index=False, name=None):
            row = []
            for v, fmt in zip(tup, fmts):
                if v is None or (not isinstance(v, str) and pd.isna(v)):
                    row.append(None)
                    continue
                wc = WriteOnlyCell(self.ws, value=v)
                if fmt:
                    wc.number_format = fmt
                row.append(wc)
            self.ws.append(row)


def _blank_repeats(df: pd.DataFrame, key: str, cols: list[str]) -> pd.DataFrame:
    dup = df.duplicated(subset=[key])
    for c in cols:
        df.loc[dup, c] = None
    return df


def _sku_pivot(df: pd.DataFrame, keys: list[str], recombine: bool) -> pd.DataFrame:
    """SKU pivot. recombine=True gives invoice semantics: rounded average
    unit price x quantity (every line internally consistent, totals drift
    from the exact books by the rounding — that drift is what the PV-sum
    checks measure). recombine=False keeps the engine's exact sums."""
    g = df.groupby(keys, as_index=False, dropna=False).agg(
        qty=("quantity", "sum"), pre_exact=("amount_pre_vat", "sum"),
        wv_exact=("amount_with_vat", "sum"))
    g["aveg"] = (g["pre_exact"] / g["qty"].replace(0, pd.NA)).fillna(0).round(0)
    factor = (g["wv_exact"] / g["pre_exact"].replace(0, pd.NA)).fillna(1.08)
    if recombine:
        g["pre"] = g["aveg"] * g["qty"]
        g["wv"] = (g["pre"] * factor).round(2)
    else:
        g["pre"] = g["pre_exact"]
        g["wv"] = g["wv_exact"]
    return g


# ---------------------------------------------------------------------------
# TikTok — tabs: PV sum, Xuat HD bt, Merries 8, KAO 8, Others 8, PV xuat HD
# ---------------------------------------------------------------------------

def build_tiktok(sku: pd.DataFrame, settings: dict, meta: dict, log: RunLog) -> tuple[Workbook, list[dict]]:
    tol = (settings.get("tolerances") or {}).get("tiktok") or {}
    tol_pv, tol_line, tol_pivot = (float(tol.get("pv_sum_vnd", 12000)),
                                   float(tol.get("xuat_hd_vnd", 2000)),
                                   float(tol.get("pv_xuat_hd_vnd", TOL_SKU_PIVOT)))
    wb = Workbook(write_only=True)
    checks: list[dict] = []
    df = sku.sort_values(["store", "order_id"]).copy()
    df["bucket"] = df["store"].map(lambda s: _bucket(s, TIKTOK_BUCKETS, "Others 8"))

    pre_total = float(df["amount_pre_vat"].sum())
    wv_total = float(df["amount_with_vat"].sum())
    by_rate_pre = {r: float(df.loc[df["vat_factor"].round(2) == r, "amount_pre_vat"].sum())
                   for r in VAT_RATES}
    recomb_wv = sum(v * r for r, v in by_rate_pre.items())

    # Brand pivots first (PV sum reads their recombined totals).
    brand_order = [("Merries 8", "Merries 8"), ("KAO 8", "KAO 8"), ("Others 8", "Others 8")]
    brand_piv = {b: _sku_pivot(df[df["bucket"] == b], ["sku_id", "sku_name"], recombine=True)
                 for _, b in brand_order}
    brand_pre = {b: float(p["pre"].sum()) for b, p in brand_piv.items()}
    brand_wv = {b: float(p["wv"].sum()) for b, p in brand_piv.items()}

    # --- PV sum ('PV sum'!A4 header; checks E2/E3 + side block F1:H8) ---
    t = _Tab(wb, "PV sum")
    pv = df.groupby(["store", "vat_factor"], as_index=False).agg(
        pre=("amount_pre_vat", "sum"), wv=("amount_with_vat", "sum"))
    side = [("Others 8%", "Others 8"), ("Kao 8%", "KAO 8"), ("Merries 8%", "Merries 8")]
    piv_pre_sum = float(sum(brand_pre.values()))
    piv_wv_sum = float(sum(brand_wv.values()))
    t.put("G1", "Sum of Amount before VAT"); t.put("H1", "Amount after VAT")
    t.put("D2", "Diff"); t.put("E2", piv_pre_sum - pre_total, ACC0)
    t.put("C3", pre_total, ACC0); t.put("D3", wv_total, ACC0)
    t.put("E3", _verdict(piv_pre_sum - pre_total, tol_pv, PVSUM_OK, PVSUM_BAD))
    for i, (label, b) in enumerate(side):
        r = 3 + i
        t.put(f"F{r}", label)
        t.put(f"G{r}", brand_pre[b], ACC0)
        t.put(f"H{r}", brand_wv[b], ACC0)
    t.put("G6", piv_pre_sum, ACC0); t.put("H6", piv_wv_sum, ACC0)
    t.put("F7", "Diff"); t.put("G7", pre_total - piv_pre_sum, ACC0)
    t.put("H7", wv_total - piv_wv_sum, ACC0)
    t.put("G8", _verdict(pre_total - piv_pre_sum, tol_pv, PVSUM_OK, PVSUM_BAD))
    t.put("H8", _verdict(wv_total - piv_wv_sum, tol_pv, PVSUM_OK, PVSUM_BAD))
    t.label_row(4, 1, ["Source.Name", "VAT KA sử dụng", "Sum of Amount before VAT", "Sum of Check total"])
    t.emit(pv[["store", "vat_factor", "pre", "wv"]], data_start_row=5,
           fmts=[None, ACC2, ACC0, ACC0], widths={"A": 34, "C": 20, "D": 20, "F": 12, "G": 18, "H": 18})
    checks.append({"tab": "PV sum", "check": "invoice pivots vs exact books (pre-VAT rounding drift)",
                   "diff": pre_total - piv_pre_sum, "tol": tol_pv,
                   "verdict": _verdict(pre_total - piv_pre_sum, tol_pv, PVSUM_OK, PVSUM_BAD)})

    # --- Xuat HD bt (header row 6; control rebuilt: exact per-VAT rows) ---
    t = _Tab(wb, "Xuat HD bt")
    out = pd.DataFrame({
        "Month": pd.to_datetime(df["income_order_created_at"]).dt.month,
        "Source.Name": df["store"],
        "Mã đơn hàng": df["order_id"],
        "Source.Name non repeat": df["store"],
        "Mã đơn hàng non repeat": df["order_id"],
        "SKU phân loại hàng": df["sku_id"],
        "Tên sản phẩm": df["sku_name"],
        "VAT KA sử dụng": df["vat_factor"],
        "Đơn giá KA sử dụng trước VAT": df["unit_price_pre_vat"],
        "số lượng KA sử dụng trước VAT": df["quantity"],
        "Amount before VAT": df["amount_pre_vat"],
        "Check total": df["amount_with_vat"],
        "Số tiền hoàn trả cho Người mua (₫)": df["actual_refund"].replace(0, pd.NA),
    })
    out = _blank_repeats(out, "Mã đơn hàng", ["Source.Name non repeat", "Mã đơn hàng non repeat"])
    t.put("L1", "check"); t.put("M1", "no VAT"); t.put("N1", "with VAT"); t.put("O1", "check")
    for i, r in enumerate(VAT_RATES):
        t.put(f"K{2 + i}", r, ACC2)
        t.put(f"M{2 + i}", by_rate_pre[r], ACC0)
        t.put(f"N{2 + i}", by_rate_pre[r] * r, ACC0)
    line_diff = wv_total - recomb_wv
    t.put("K5", "total"); t.put("L5", wv_total, ACC0); t.put("N5", recomb_wv, ACC0)
    t.put("O5", line_diff, ACC0); t.put("P5", _verdict(line_diff, tol_line))
    t.label_row(6, 1, list(out.columns))
    t.emit(out, data_start_row=7,
           fmts=[None, None, None, None, None, None, None, ACC2, ACC0, ACC0, ACC0, ACC0, ACC0],
           widths={"B": 26, "C": 22, "G": 44, "I": 16, "K": 16, "L": 16, "M": 16},
           freeze="A7")
    checks.append({"tab": "Xuat HD bt", "check": "with-VAT lines vs exact VAT recombination",
                   "diff": line_diff, "tol": tol_line, "verdict": _verdict(line_diff, tol_line)})

    # --- Brand tabs (header row 4; totals at E2/F2 like the template) ---
    for tab_name, b in brand_order:
        t = _Tab(wb, tab_name)
        piv = brand_piv[b]
        sub = df[df["bucket"] == b]
        stores = sub["store"].unique()
        t.put("A1", "VAT KA sử dụng")
        t.put("B1", float(sub["vat_factor"].max()) if len(sub) else None, ACC2)
        t.put("A2", "Source.Name")
        t.put("B2", stores[0] if len(stores) == 1 else "(Multiple Items)")
        t.put("E2", brand_pre[b], ACC0)
        t.put("F2", brand_wv[b], ACC0)
        t.label_row(4, 1, ["SKU phân loại hàng", "Tên sản phẩm", "Sum of số lượng KA sử dụng trước VAT",
                           "Sum of Aveg Price KA sử dụng", "Sum of Check Amount No VAT", "Sum of Check total"])
        t.emit(piv[["sku_id", "sku_name", "qty", "aveg", "pre", "wv"]], data_start_row=5,
               fmts=[None, None, ACC0, ACC0, ACC0, ACC0],
               widths={"A": 22, "B": 46, "C": 14, "D": 16, "E": 18, "F": 18})

    # --- PV xuat HD (exact SKU pivot; drop-detection check at 1,000) ---
    t = _Tab(wb, "PV xuat HD")
    piv = _sku_pivot(df, ["store", "sku_id", "sku_name"], recombine=False)
    pivot_pre = float(piv["pre"].sum())
    t.put("A1", "VAT KA sử dụng"); t.put("B1", "(All)")
    t.put("E1", "Sale source data"); t.put("F1", pre_total, ACC0)
    t.put("G1", "Diff"); t.put("H1", pre_total - pivot_pre, ACC0)
    t.put("I1", _verdict(pre_total - pivot_pre, tol_pivot))
    t.put("F2", pivot_pre, ACC0); t.put("G2", float(piv["wv"].sum()), ACC0)
    t.label_row(3, 1, ["Source.Name", "SKU phân loại hàng", "Tên sản phẩm",
                       "Sum of số lượng KA sử dụng trước VAT", "Sum of Aveg Price KA sử dụng",
                       "Sum of Check Amount No VAT", "Sum of Check total"])
    t.emit(piv[["store", "sku_id", "sku_name", "qty", "aveg", "pre", "wv"]], data_start_row=4,
           fmts=[None, None, None, ACC0, ACC0, ACC0, ACC0],
           widths={"A": 30, "B": 22, "C": 46, "F": 18, "G": 18})
    checks.append({"tab": "PV xuat HD", "check": "pre-VAT lines vs exact SKU pivot",
                   "diff": pre_total - pivot_pre, "tol": tol_pivot,
                   "verdict": _verdict(pre_total - pivot_pre, tol_pivot)})
    return wb, checks


# ---------------------------------------------------------------------------
# Shopee — tabs: PV sum, Xuat HD bt, return, 1.05 KA, Curel, KAO, Merries,
#                Kate, Others, 1.08 KA, 1.10 KA, PV xuat HD (template order)
# ---------------------------------------------------------------------------

def build_shopee(sku: pd.DataFrame, settings: dict, meta: dict, log: RunLog) -> tuple[Workbook, list[dict]]:
    tol = (settings.get("tolerances") or {}).get("shopee") or {}
    tol_line = float(tol.get("xuat_hd_vnd", 2000))
    tol_ret = float(tol.get("return_full_vnd", 10))
    tol_drift = TOL_PIVOT_DRIFT_SHOPEE
    wb = Workbook(write_only=True)
    checks: list[dict] = []

    df = sku.sort_values(["store", "order_id"]).copy()
    ok = df[df["check_status"] == "ok"].copy()
    ret = df[df["check_status"] == "Return"].copy()
    ok["bucket"] = ok["store"].map(lambda s: _bucket(s, SHOPEE_BUCKETS, "Others"))

    pre_total = float(ok["amount_pre_vat"].sum())
    wv_total = float(ok["amount_with_vat"].sum())
    by_rate_pre = {r: float(ok.loc[ok["vat_factor"].round(2) == r, "amount_pre_vat"].sum())
                   for r in VAT_RATES}
    recomb_wv = sum(v * r for r, v in by_rate_pre.items())

    # Pivot frames first (PV sum reads their recombined totals).
    brand_names = ("Curel", "KAO", "Merries", "Kate", "Others")
    brand_piv = {b: _sku_pivot(ok[ok["bucket"] == b], ["sku_id", "sku_name"], recombine=True)
                 for b in brand_names}
    brand_pre = {b: float(p["pre"].sum()) for b, p in brand_piv.items()}
    vat_piv = {r: _sku_pivot(ok[ok["vat_factor"].round(2) == r],
                             ["store", "sku_id", "sku_name"], recombine=True)
               for r in VAT_RATES}
    vat_pre = {r: float(p["pre"].sum()) for r, p in vat_piv.items()}

    # Partial-return invoiceable total = the template's 'return'!Q5.
    per_order = ret.groupby("order_id", as_index=False).agg(
        total=("amount_with_vat", "sum"), refund=("actual_refund", "max"))
    per_order["refund_adj"] = per_order["refund"].fillna(0).map(lambda v: -v if v > 0 else v)
    per_order["check"] = per_order["total"] + per_order["refund_adj"]
    partial_total = float(per_order["check"].sum())

    # --- PV sum (verdict wired to the real diff; template's G3 read blank E2) ---
    t = _Tab(wb, "PV sum")
    pv = ok.copy()
    pv["com"] = pd.to_datetime(pv["income_order_created_at"]).dt.month
    pv["fm"] = pd.to_datetime(pv["statement_date"]).dt.month
    pvt = pv.groupby(["store", "com", "fm", "vat_factor"], as_index=False).agg(
        pre=("amount_pre_vat", "sum"), wv=("amount_with_vat", "sum"))
    side_rows = ([("5", None, vat_pre[1.05], 1.05)]
                 + [("8", b, brand_pre[b], 1.08) for b in ("Others", "Curel", "KAO", "Merries", "Kate")]
                 + [("10", None, vat_pre[1.10], 1.10)])
    j9 = sum(v for _, _, v, _ in side_rows)
    k9 = sum(v * f for _, _, v, f in side_rows)
    t.put("J1", "Sum of Amount before VAT"); t.put("K1", "Amount after VAT")
    t.put("F2", "Diff"); t.put("G2", j9 - pre_total, ACC0)
    t.put("E3", pre_total, ACC0); t.put("F3", wv_total, ACC0)
    t.put("G3", _verdict(j9 - pre_total, tol_drift, PVSUM_OK, PVSUM_BAD))
    for i, (rate_lab, name, val, factor) in enumerate(side_rows):
        r = 2 + i
        t.put(f"H{r}", rate_lab)
        if name:
            t.put(f"I{r}", name)
        t.put(f"J{r}", val, ACC0)
        t.put(f"K{r}", val * factor, ACC0)
    t.put("J9", j9, ACC0); t.put("K9", k9, ACC0)
    t.put("I10", "Diff"); t.put("J10", pre_total - j9, ACC0); t.put("K10", wv_total - k9, ACC0)
    t.put("J11", _verdict(pre_total - j9, tol_drift, PVSUM_OK, PVSUM_BAD))
    t.put("K11", _verdict(wv_total - k9, tol_drift, PVSUM_OK, PVSUM_BAD))
    t.label_row(4, 1, ["Source.Name", "Create_Order_Month", "Finance_Month", "VAT KA sử dụng",
                       "Sum of Amount before VAT", "Sum of Check total"])
    t.emit(pvt[["store", "com", "fm", "vat_factor", "pre", "wv"]], data_start_row=5,
           fmts=[None, None, None, ACC2, ACC0, ACC0],
           widths={"A": 34, "E": 20, "F": 20, "I": 10, "J": 18, "K": 18})
    checks.append({"tab": "PV sum", "check": "invoice pivots vs exact books (pre-VAT rounding drift)",
                   "diff": pre_total - j9, "tol": tol_drift,
                   "verdict": _verdict(pre_total - j9, tol_drift, PVSUM_OK, PVSUM_BAD)})

    # --- Xuat HD bt (15 cols incl refund; control block exact) ---
    t = _Tab(wb, "Xuat HD bt")
    out = pd.DataFrame({
        "Diff": ok["check_status"],
        "Source.Name": ok["store"],
        "Mã đơn hàng": ok["order_id"],
        "Source.Name non repeat": ok["store"],
        "Mã đơn hàng non repeat": ok["order_id"],
        "Create_Order_Month": pd.to_datetime(ok["income_order_created_at"]).dt.month,
        "Finance_Month": pd.to_datetime(ok["statement_date"]).dt.month,
        "SKU phân loại hàng": ok["sku_id"],
        "Tên sản phẩm": ok["sku_name"],
        "VAT KA sử dụng": ok["vat_factor"],
        "Đơn giá KA sử dụng trước VAT": ok["unit_price_pre_vat"],
        "số lượng KA sử dụng trước VAT": ok["quantity"],
        "Amount before VAT": ok["amount_pre_vat"],
        "Check total": ok["amount_with_vat"],
        "Số tiền hoàn trả cho Người mua (₫)": ok["actual_refund"].replace(0, pd.NA),
    })
    out = _blank_repeats(out, "Mã đơn hàng", ["Source.Name non repeat", "Mã đơn hàng non repeat"])
    t.put("N1", "check"); t.put("O1", "no VAT"); t.put("P1", "with VAT"); t.put("Q1", "check")
    t.put("M2", "Tổng xuất HD (lines + return 1 phần)"); t.put("N2", wv_total + partial_total, ACC0)
    t.put("M3", "Return 1 phần phải xuất HD"); t.put("N3", partial_total, ACC0)
    for i, r in enumerate(VAT_RATES):
        t.put(f"L{2 + i}", r, ACC2)
        t.put(f"O{2 + i}", by_rate_pre[r], ACC0)
        t.put(f"P{2 + i}", by_rate_pre[r] * r, ACC0)
    line_diff = wv_total - recomb_wv
    t.put("M5", "total"); t.put("N5", wv_total, ACC0); t.put("P5", recomb_wv, ACC0)
    t.put("Q5", line_diff, ACC0); t.put("R5", _verdict(line_diff, tol_line))
    t.label_row(6, 1, list(out.columns))
    t.emit(out, data_start_row=7,
           fmts=[None, None, None, None, None, None, None, None, None, ACC2, ACC0, ACC0, ACC0, ACC0, ACC0],
           widths={"B": 26, "C": 20, "I": 44, "K": 16, "M": 16, "N": 16}, freeze="A7")
    checks.append({"tab": "Xuat HD bt", "check": "with-VAT lines vs exact VAT recombination",
                   "diff": line_diff, "tol": tol_line, "verdict": _verdict(line_diff, tol_line)})

    # --- return (template geometry starts at col C with an unlabeled check
    #     column; cleaned: anchored at A, suffixed repeat headers, 'Check'
    #     labeled — column CONTENT and order unchanged) ---
    t = _Tab(wb, "return")
    rtab = pd.DataFrame({
        "Diff": ret["check_status"],
        "Source.Name": ret["store"],
        "Mã đơn hàng": ret["order_id"],
        "Source.Name non repeat": ret["store"],
        "Mã đơn hàng non repeat": ret["order_id"],
        "SKU phân loại hàng": ret["sku_id"],
        "Tên sản phẩm": ret["sku_name"],
        "VAT KA sử dụng": ret["vat_factor"],
        "Đơn giá KA sử dụng trước VAT": ret["unit_price_pre_vat"],
        "số lượng KA sử dụng trước VAT": ret["quantity"],
        "Cộng tiền hàng KA sử dụng trước VAT": ret["amount_pre_vat"],
        "Cộng tiền hàng KA sử dụng có VAT": ret["amount_with_vat"],
    })
    order_map = per_order.set_index("order_id")
    rtab["Total by order"] = ret["order_id"].map(order_map["total"]).values
    rtab["Số tiền hoàn trả cho Người mua (₫)"] = ret["order_id"].map(order_map["refund_adj"]).values
    rtab["Check"] = ret["order_id"].map(order_map["check"]).values
    rtab["Note"] = rtab["Check"].map(lambda v: RETURN_FULL if abs(v) < tol_ret else RETURN_PARTIAL)
    rtab = _blank_repeats(rtab, "Mã đơn hàng",
                          ["Source.Name non repeat", "Mã đơn hàng non repeat",
                           "Total by order", "Số tiền hoàn trả cho Người mua (₫)", "Check", "Note"])
    t.put("B2", "Return bang dung so tien mua")
    t.put("M5", float(per_order["total"].sum()), ACC0)
    t.put("N5", float(per_order["refund_adj"].sum()), ACC0)
    t.put("O5", partial_total, ACC0)
    t.label_row(6, 1, list(rtab.columns))
    t.emit(rtab, data_start_row=7,
           fmts=[None, None, None, None, None, None, None, ACC2, ACC0, ACC0, ACC0, ACC0, ACC0, ACC0, ACC0, None],
           widths={"B": 26, "C": 20, "G": 44, "M": 16, "N": 16, "O": 14, "P": 26}, freeze="A7")
    n_partial = int((per_order["check"].abs() >= tol_ret).sum())
    checks.append({"tab": "return", "check": f"partial returns to invoice: {n_partial} orders",
                   "diff": partial_total, "tol": tol_ret, "verdict": f"{partial_total:,.0f} VND"})

    # --- brand + VAT pivot tabs (recombined), in the template's tab order ---
    def brand_tab(name: str) -> None:
        t = _Tab(wb, name)
        piv = brand_piv[name]
        sub = ok[ok["bucket"] == name]
        stores = sub["store"].unique()
        t.put("A1", "VAT KA sử dụng")
        t.put("B1", float(sub["vat_factor"].max()) if len(sub) else None, ACC2)
        t.put("A2", "Source.Name"); t.put("B2", stores[0] if len(stores) == 1 else "(Multiple Items)")
        t.put("E3", brand_pre[name], ACC0)   # template total position: brand!E3
        t.label_row(4, 1, ["SKU phân loại hàng", "Tên sản phẩm", "Sum of số lượng KA sử dụng trước VAT",
                           "Sum of Aveg Price KA sử dụng", "Sum of Check Amount No VAT"])
        t.emit(piv[["sku_id", "sku_name", "qty", "aveg", "pre"]], data_start_row=5,
               fmts=[None, None, ACC0, ACC0, ACC0], widths={"A": 22, "B": 46, "E": 18})

    def vat_tab(name: str, rate: float) -> None:
        t = _Tab(wb, name)
        piv = vat_piv[rate]
        t.put("A1", "VAT KA sử dụng"); t.put("B1", rate if len(piv) else None, ACC2)
        t.put("F2", vat_pre[rate], ACC0)     # template total position: ' X KA'!F2
        t.label_row(3, 1, ["Source.Name", "SKU phân loại hàng", "Tên sản phẩm",
                           "Sum of số lượng KA sử dụng trước VAT", "Sum of Aveg Price KA sử dụng",
                           "Sum of Check Amount No VAT"])
        t.emit(piv[["store", "sku_id", "sku_name", "qty", "aveg", "pre"]], data_start_row=4,
               fmts=[None, None, None, ACC0, ACC0, ACC0], widths={"A": 30, "B": 22, "C": 46, "F": 18})

    vat_tab("1.05 KA", 1.05)
    for b in brand_names:
        brand_tab(b)
    vat_tab("1.08 KA", 1.08)
    vat_tab("1.10 KA", 1.10)

    # --- PV xuat HD (exact SKU pivot; drop-detection check at 1,000) ---
    t = _Tab(wb, "PV xuat HD")
    piv = _sku_pivot(ok, ["store", "sku_id", "sku_name"], recombine=False)
    pivot_pre = float(piv["pre"].sum())
    t.put("A1", "VAT KA sử dụng"); t.put("B1", "(All)")
    t.put("E1", "Sale source data"); t.put("F1", pre_total, ACC0)
    t.put("G1", "Diff"); t.put("H1", pre_total - pivot_pre, ACC0)
    t.put("I1", _verdict(pre_total - pivot_pre, TOL_SKU_PIVOT))
    t.put("E2", pivot_pre, ACC0); t.put("F2", float(piv["wv"].sum()), ACC0)
    t.label_row(3, 1, ["Source.Name", "SKU phân loại hàng", "Tên sản phẩm",
                       "Sum of số lượng KA sử dụng trước VAT", "Sum of Amount before VAT",
                       "Sum of Check total"])
    t.emit(piv[["store", "sku_id", "sku_name", "qty", "pre", "wv"]], data_start_row=4,
           fmts=[None, None, None, ACC0, ACC0, ACC0], widths={"A": 30, "B": 22, "C": 46, "E": 18, "F": 18})
    checks.append({"tab": "PV xuat HD", "check": "pre-VAT lines vs exact SKU pivot",
                   "diff": pre_total - pivot_pre, "tol": TOL_SKU_PIVOT,
                   "verdict": _verdict(pre_total - pivot_pre, TOL_SKU_PIVOT)})
    return wb, checks


# ---------------------------------------------------------------------------
# Lazada — tabs (cleaned order): Summary, then the 1.05/1.08/1.10 pairs
# ('X KA used' + 'X'), then brand tabs. The template scattered these.
# ---------------------------------------------------------------------------

def build_lazada(rev: pd.DataFrame, settings: dict, meta: dict, log: RunLog,
                 classified: pd.DataFrame | None = None) -> tuple[Workbook, list[dict]]:
    wb = Workbook(write_only=True)
    checks: list[dict] = []
    month_label = meta.get("month_label", "")
    period_label = meta.get("period_label", "")

    df = rev.sort_values(["store", "order_id"]).copy()
    df["bucket"] = df["store"].map(lambda s: _bucket(s, LAZADA_BUCKETS, "Others"))
    df["money"] = df["credits"].fillna(0) + df["promo"].fillna(0)   # actual settled VND incl VAT

    # Order-less revenue: ledger rows the team's Lib maps to 1.Doanh Thu but
    # that carry no order/SKU/quantity (July 2026: "Lost & Damaged FBL
    # Inventories" compensation). They are real settled money — kept in the
    # sale-report figure — but can never be invoice lines, so they appear as
    # a NAMED reconciling row in each control block instead of silently
    # distorting (old plain export) or failing (naive check) it.
    orderless_mask = (df["quantity"].fillna(0) <= 0) | df["order_id"].isin(["0", "", "nan", "None"])
    orderless = df[orderless_mask]
    if len(orderless):
        for _, r in orderless.iterrows():
            log.warn(f"order-less revenue kept as reconciling item, NOT an invoice line: "
                     f"{r['store']} {r['money']:,.0f} VND @ {r['vat_rate']}")
    df = df[~orderless_mask]

    def laz_pivot(sub: pd.DataFrame) -> pd.DataFrame:
        """Invoice-view SKU pivot: rounded average price x qty (per-SKU rate)."""
        g = sub.groupby(["sku_id", "product_name"], as_index=False, dropna=False).agg(
            qty=("quantity", "sum"), no_vat_exact=("check_no_vat", "sum"),
            rate=("vat_rate", "max"))
        g["aveg"] = (g["no_vat_exact"] / g["qty"].replace(0, pd.NA)).fillna(0).round(0)
        g["no_vat"] = g["aveg"] * g["qty"]
        g["wv"] = (g["no_vat"] * g["rate"]).round(2)
        return g

    def at(rate: float) -> pd.DataFrame:
        return df[df["vat_rate"].round(2) == rate]

    # Sale report = ALL Doanh Thu money incl. order-less items (matching the
    # team's bucket view); the invoiceable portion excludes them.
    orderless_rate = {r: float(orderless.loc[orderless["vat_rate"].round(2) == r, "money"].sum())
                      for r in VAT_RATES}
    sale_report = {r: float(at(r)["money"].sum()) + orderless_rate[r] for r in VAT_RATES}
    no_vat_exact = {r: float(at(r)["check_no_vat"].sum()) for r in VAT_RATES}
    rate_piv = {r: laz_pivot(at(r)) for r in VAT_RATES}
    brand_piv = {b: laz_pivot(df[df["bucket"] == b])
                 for b in ("Curel.xlsx", "KAO.xlsx", "Merries.xlsx", "Others")}
    brand_wv = {b: float(p["wv"].sum()) for b, p in brand_piv.items()}
    rate_wv = {r: float(p["wv"].sum()) for r, p in rate_piv.items()}
    rate_novat = {r: float(p["no_vat"].sum()) for r, p in rate_piv.items()}

    # --- Summary (template row positions D6:F17) ---
    t = _Tab(wb, "Summary")
    t.label_row(6, 4, ["Type", "VAT rate KA use", "Total order"])
    t.put("E7", 1.05, ACC2); t.put("F7", rate_wv[1.05], PLAIN2)
    for i, b in enumerate(["Others", "Curel.xlsx", "KAO.xlsx", "Merries.xlsx"]):
        r = 8 + i
        t.put(f"D{r}", b.replace(".xlsx", ""))
        t.put(f"E{r}", 1.08, ACC2)
        t.put(f"F{r}", brand_wv[b], PLAIN2)
    t.put("E12", 1.10, ACC2); t.put("F12", rate_wv[1.10], PLAIN2)
    total_wv = rate_wv[1.05] + float(sum(brand_wv.values())) + rate_wv[1.10]
    report_total = sum(sale_report.values())
    orderless_total = float(sum(orderless_rate.values()))
    prior_invoiced = 0.0   # "Order đã xuất HD trước đó" — manual field, defaults to 0
    summary_diff = report_total - orderless_total - prior_invoiced - total_wv
    t.put("D13", "Total"); t.put("F13", total_wv, PLAIN2)
    t.put("D14", "Per finance report"); t.put("F14", report_total, PLAIN2)
    t.put("D15", "Doanh thu không theo đơn hàng (không xuất HD dòng)")
    t.put("F15", orderless_total, PLAIN2)
    t.put("D16", "Order đã xuất HD trước đó"); t.put("F16", prior_invoiced, PLAIN2)
    t.put("D17", "Diff"); t.put("F17", summary_diff, PLAIN2)
    t.put("D18", "Check"); t.put("F18", _verdict(summary_diff, TOL_PIVOT_DRIFT_LAZADA))
    t.emit(widths={"D": 44, "E": 16, "F": 20})
    checks.append({"tab": "Summary", "check": "sale report vs invoice total (with VAT)",
                   "diff": summary_diff, "tol": TOL_PIVOT_DRIFT_LAZADA,
                   "verdict": _verdict(summary_diff, TOL_PIVOT_DRIFT_LAZADA)})

    # --- per-rate pair: 'X KA used' (SKU invoice view) + 'X' (exact lines) ---
    for rate in VAT_RATES:
        key = {1.05: "1.05", 1.08: "1.08", 1.10: "1.10"}[rate]
        sub = at(rate)
        piv = rate_piv[rate]

        t = _Tab(wb, f"{key} KA used")
        diff = rate_wv[rate] - (sale_report[rate] - orderless_rate[rate])
        t.put("A1", "Type"); t.put("B1", rate, ACC2)
        t.put("C1", "Month"); t.put("D1", month_label)
        t.put("F1", "số trên sale report"); t.put("G1", sale_report[rate], PLAIN2)
        t.put("H1", "Diff"); t.put("I1", diff, PLAIN2)
        t.put("J1", _verdict(diff, TOL_PIVOT_DRIFT_LAZADA))
        t.put("A2", "Period"); t.put("B2", period_label)
        if orderless_rate[rate]:
            t.put("F2", "trong đó: không theo đơn hàng")
            t.put("G2", orderless_rate[rate], PLAIN2)
        t.label_row(4, 1, ["KA Sử dụng cột này"] * 4 + ["check only", "check only"])
        t.put("E5", rate_novat[rate], PLAIN0); t.put("F5", rate_wv[rate], PLAIN2)
        t.label_row(6, 1, ["Seller SKU", "Details", "Sum of Sum of Quantity",
                           "Sum of Average Price KA used", "Sum of Check average amount",
                           "Sum of  check avg amount with VAT"])
        t.emit(piv[["sku_id", "product_name", "qty", "aveg", "no_vat", "wv"]], data_start_row=7,
               fmts=[None, None, PLAIN0, PLAIN0, PLAIN0, PLAIN2],
               widths={"A": 22, "B": 50, "D": 18, "E": 18, "F": 20})
        checks.append({"tab": f"{key} KA used", "check": "invoice pivot with-VAT vs sale report",
                       "diff": diff, "tol": TOL_PIVOT_DRIFT_LAZADA,
                       "verdict": _verdict(diff, TOL_PIVOT_DRIFT_LAZADA)})

        t = _Tab(wb, key)
        lines = pd.DataFrame({
            "Source.Name": sub["store"],
            "Order No.": sub["order_id"],
            "Seller SKU": sub["sku_id"],
            "Details": sub["product_name"],
            "Sum of Price KA": sub["price_ka"],
            "Sum of Quantity": sub["quantity"],
            "Đơn Vị Tính": None,
            "Amount": sub["check_no_vat"],
            "Xuất HD khách": None,
            "Note": None,
        })
        vat_amount = no_vat_exact[rate] * rate
        line_diff = sale_report[rate] - orderless_rate[rate] - vat_amount
        t.put("A1", "Type"); t.put("B1", rate, ACC2)
        t.put("C1", "Month"); t.put("D1", month_label)
        t.put("A2", "Period"); t.put("B2", period_label)
        t.put("G2", "Diff"); t.put("H2", line_diff, PLAIN2)
        t.put("I2", _verdict(line_diff, TOL_LAZ_LINE))
        t.put("G3", "Số trên sale report"); t.put("H3", sale_report[rate], PLAIN2)
        t.put("G4", "VAT"); t.put("H4", vat_amount, PLAIN2)
        t.put("G5", "no Vat"); t.put("H5", no_vat_exact[rate], PLAIN0)
        if orderless_rate[rate]:
            t.put("G6", "Không theo đơn hàng (không xuất HD dòng)")
            t.put("H6", orderless_rate[rate], PLAIN2)
        t.label_row(6, 1, list(lines.columns))
        t.emit(lines, data_start_row=7,
               fmts=[None, None, None, None, PLAIN0, PLAIN0, None, ACC0, None, None],
               widths={"A": 24, "B": 20, "C": 22, "D": 50, "E": 14, "H": 16}, freeze="A7")
        checks.append({"tab": key, "check": "sale report vs exact lines x VAT",
                       "diff": line_diff, "tol": TOL_LAZ_LINE,
                       "verdict": _verdict(line_diff, TOL_LAZ_LINE)})

    # --- brand tabs (template: totals at F2, header row 3) ---
    for b in ("Curel.xlsx", "KAO.xlsx", "Merries.xlsx", "Others"):
        t = _Tab(wb, b)
        piv = brand_piv[b]
        stores = df.loc[df["bucket"] == b, "store"].unique()
        t.put("A1", "Source.Name"); t.put("B1", stores[0] if len(stores) == 1 else "(Multiple Items)")
        t.put("F2", brand_wv[b], PLAIN2)
        t.label_row(3, 1, ["Seller SKU", "Details", "Sum of Sum of Quantity",
                           "Sum of Average Price KA used", "Sum of Check average amount",
                           "Sum of  check avg amount with VAT"])
        t.emit(piv[["sku_id", "product_name", "qty", "aveg", "no_vat", "wv"]], data_start_row=4,
               fmts=[None, None, PLAIN0, PLAIN0, PLAIN0, PLAIN2],
               widths={"A": 22, "B": 50, "E": 18, "F": 20})

    # Extra tab beyond the team's template: the fee-bucket overview the
    # plain export used to carry (the team's own fee view lives in their
    # Total file's SUM CP, which the invoicing template does not include).
    if classified is not None and len(classified):
        t = _Tab(wb, "Fee buckets")
        buckets = (classified.pivot_table(index="store", columns="fee_bucket",
                                          values="amount_incl_vat", aggfunc="sum")
                   .round(2).reset_index())
        t.label_row(1, 1, [str(c) for c in buckets.columns])
        t.emit(buckets, data_start_row=2,
               fmts=[None] + [PLAIN2] * (buckets.shape[1] - 1),
               widths={"A": 30})
    return wb, checks


def write_workbook(wb: Workbook, path: Path, checks: list[dict], log: RunLog) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    log.add(f"  wrote {path.name} ({len(wb.sheetnames)} tabs: {wb.sheetnames})")
    for c in checks:
        log.add(f"    check [{c['tab']}] {c['check']}: diff {c['diff']:,.2f} "
                f"(tol {c['tol']:,.0f}) -> {c['verdict']}")
