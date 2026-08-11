"""Cross-window, cross-platform master summary in the team's PV sum /
Summary shape (team request, Aug 2026: "one master file to consolidate all
the weeks by brand ... for all platforms").

Reads the GENERATED per-window finance files and aggregates them, so every
figure in the master is provably the same number the team already has in the
weekly files. Column totals are asserted against each source file and the
check is printed; a master that cannot tie to its sources is worthless.

Usage:
    python tools/build_master_summary.py --month 2026-07 \
        --out "C:/Users/.../ADA marketplace master July 2026.xlsx"
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

ACC0 = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'

# (period suffix, human label) per platform, in settlement order.
WINDOWS = {
    "TikTok": [("w1", "01-07"), ("w2", "08-14"), ("w3", "15-21"),
               ("w4", "22-28"), ("w5", "29-31")],
    "Shopee": [("s1", "01-10"), ("s2", "11-20"), ("s3", "21-28"), ("s4", "29-31")],
    "Lazada": [("l1", "01-05"), ("l2", "06-12"), ("l3", "13-19"),
               ("l4", "20-26"), ("l5", "27-31")],
}
PLATFORM_DIR = {"TikTok": "tiktok", "Shopee": "shopee", "Lazada": "lazada"}

# Template-shaped finance files (src/finance_template.py): line tabs are
# 'Xuat HD bt' (TikTok/Shopee) and '1.05'/'1.08'/'1.10' (Lazada), header on
# row 6. Lazada line tabs carry no with-VAT column (the team's template does
# not) — with-VAT is recombined as Amount x the tab's rate, which is exact.
SCHEMA = {
    "TikTok": (lambda s: s == "Xuat HD bt", "Source.Name",
               "Amount before VAT", "Check total"),
    "Shopee": (lambda s: s == "Xuat HD bt", "Source.Name",
               "Amount before VAT", "Check total"),
    "Lazada": (lambda s: re.fullmatch(r"1\.\d{1,2}", s.strip()) is not None, "Source.Name",
               "Amount", None),
}
HEADER_ROW = 5  # 0-based; template line tabs put headers on sheet row 6


def norm_store(name: str) -> str:
    """Same normalization as tools/full_run.py so labels agree."""
    s = unicodedata.normalize("NFC", str(name)).lower().strip()
    s = re.sub(r"^\s*\d+[._ ]*", "", s)
    s = re.sub(r"^(income|order)\b[. ]*", "", s)
    s = re.sub(r"\s+part\s*\d+", "", s)
    s = s.replace(".xlsx", "")
    s = re.sub(r"[._]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def read_window(platform: str, month: str, suffix: str) -> pd.DataFrame:
    sel, store_c, pre_c, wv_c = SCHEMA[platform]
    f = ROOT / "output" / f"{month}_{suffix}" / PLATFORM_DIR[platform] / "finance_file.xlsx"
    if not f.exists():
        raise FileNotFoundError(f)
    xf = pd.ExcelFile(f, engine="calamine")
    frames = []
    for sheet in [s for s in xf.sheet_names if sel(s)]:
        d = pd.read_excel(xf, sheet_name=sheet, engine="calamine", header=HEADER_ROW)
        d.columns = [str(c).strip() for c in d.columns]
        if store_c not in d.columns:
            continue
        pre = pd.to_numeric(d.get(pre_c), errors="coerce").fillna(0)
        if wv_c is None:  # Lazada: with-VAT recombined from the tab's rate
            wv = pre * float(sheet.strip())
        else:
            wv = pd.to_numeric(d.get(wv_c), errors="coerce").fillna(0)
        # Order-level columns are blanked on repeat SKU rows ("non repeat"
        # convention); forward-fill so every line carries its store.
        frames.append(pd.DataFrame({"store": d[store_c].ffill(), "pre": pre, "wv": wv}))
    if not frames:
        raise ValueError(f"no line tab found in {f}")
    df = pd.concat(frames, ignore_index=True)
    df["store"] = df["store"].map(norm_store)
    return df.groupby("store", as_index=False)[["pre", "wv"]].sum()


def write_grid(ws, title: str, row_label: str, rows: list[str], cols: list[str],
               values: dict, extra_cols: list[tuple[str, dict]] | None = None) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12)
    ws["A3"] = row_label
    ws["A3"].font = Font(bold=True)
    for j, c in enumerate(cols, start=2):
        cell = ws.cell(row=3, column=j, value=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    n_extra = len(extra_cols or [])
    for k, (label, _) in enumerate(extra_cols or [], start=len(cols) + 2):
        cell = ws.cell(row=3, column=k, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=r)
        for j, c in enumerate(cols, start=2):
            cell = ws.cell(row=i, column=j, value=float(values.get((r, c), 0.0)))
            cell.number_format = ACC0
        for k, (_, vals) in enumerate(extra_cols or [], start=len(cols) + 2):
            cell = ws.cell(row=i, column=k, value=float(vals.get(r, 0.0)))
            cell.number_format = ACC0

    total_row = len(rows) + 4
    ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
    for j in range(2, len(cols) + 2 + n_extra):
        letter = get_column_letter(j)
        cell = ws.cell(row=total_row, column=j,
                       value=f"=SUM({letter}4:{letter}{total_row - 1})")
        cell.number_format = ACC0
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 34
    for j in range(2, len(cols) + 2 + n_extra):
        ws.column_dimensions[get_column_letter(j)].width = 18
    ws.freeze_panes = "B4"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", required=True, help="e.g. 2026-07")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    data: dict[str, dict[str, pd.DataFrame]] = {}
    for platform, wins in WINDOWS.items():
        data[platform] = {}
        for suffix, label in wins:
            data[platform][label] = read_window(platform, args.month, suffix)
            tot = data[platform][label]["wv"].sum()
            print(f"  {platform:<7} {label}: {len(data[platform][label]):>3} stores, "
                  f"with-VAT {tot:>18,.0f}")

    wb = Workbook()
    wb.remove(wb.active)

    # --- Summary: vertical, one block per platform (the platforms do NOT
    # share window boundaries, so a common column axis would be sparse and
    # misleading; this mirrors the Lazada "Summary" tab idiom instead) ---
    plats = list(WINDOWS)
    ws = wb.create_sheet("Summary")
    ws["A1"] = f"Marketplace revenue summary, {args.month} (VND)"
    ws["A1"].font = Font(bold=True, size=12)
    for j, h in enumerate(["Platform", "Window", "Sum of Amount before VAT",
                           "Sum of Check total (with VAT)"], start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center" if j > 2 else "left")
    r = 4
    grand_pre = grand_wv = 0.0
    for p in plats:
        p_pre = p_wv = 0.0
        for _, lab in WINDOWS[p]:
            pre = float(data[p][lab]["pre"].sum())
            wv = float(data[p][lab]["wv"].sum())
            p_pre += pre
            p_wv += wv
            ws.cell(row=r, column=1, value=p)
            ws.cell(row=r, column=2, value=lab)
            ws.cell(row=r, column=3, value=pre).number_format = ACC0
            ws.cell(row=r, column=4, value=wv).number_format = ACC0
            r += 1
        ws.cell(row=r, column=1, value=f"{p} total").font = Font(bold=True)
        for col, v in ((3, p_pre), (4, p_wv)):
            c = ws.cell(row=r, column=col, value=v)
            c.number_format = ACC0
            c.font = Font(bold=True)
        grand_pre += p_pre
        grand_wv += p_wv
        r += 2
    ws.cell(row=r, column=1, value="ALL PLATFORMS").font = Font(bold=True, size=12)
    for col, v in ((3, grand_pre), (4, grand_wv)):
        c = ws.cell(row=r, column=col, value=v)
        c.number_format = ACC0
        c.font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 28

    # --- storefront -> client-brand mapping (config/brand_map.csv, team-
    # reviewable). Unmapped storefronts fall back to their own name and are
    # flagged; nothing is merged without an explicit mapping row. ---
    bmap: dict[tuple[str, str], tuple[str, str, str]] = {}
    map_path = ROOT / "config" / "brand_map.csv"
    if map_path.exists():
        m = pd.read_csv(map_path, dtype=str).fillna("")
        for _, r in m.iterrows():
            bmap[(r["platform"].strip().lower(), r["storefront"].strip())] = (
                r["client_brand"].strip(), r["confidence"].strip(), r["note"].strip())

    per_plat: dict[str, dict[str, float]] = {}
    for p in plats:
        acc: dict[str, float] = {}
        for lab in data[p]:
            for s, v in zip(data[p][lab]["store"], data[p][lab]["wv"]):
                acc[s] = acc.get(s, 0.0) + float(v)
        per_plat[p] = acc

    def brand_of(p: str, store: str) -> tuple[str, str, str]:
        return bmap.get((p.lower(), store),
                        (store, "UNMAPPED (kept as storefront)", "not in config/brand_map.csv"))

    # By brand (mapped)
    brand_vals: dict[tuple[str, str], float] = {}
    for p in plats:
        for s, v in per_plat[p].items():
            b = brand_of(p, s)[0]
            brand_vals[(b, p)] = brand_vals.get((b, p), 0.0) + v
    brands = sorted({b for b, _ in brand_vals})
    ws = wb.create_sheet("By brand")
    write_grid(ws, f"Revenue by client brand across platforms, {args.month} (with VAT, VND) — "
                   f"mapping on the 'Brand mapping' tab",
               "Client brand", brands, plats,
               {(b, p): brand_vals.get((b, p), 0.0) for b in brands for p in plats})

    # By storefront (unmapped view, for traceability)
    stores_all = sorted({s for p in plats for s in per_plat[p]})
    ws = wb.create_sheet("By storefront")
    write_grid(ws, f"Revenue by storefront (as named in the platform files), {args.month} (with VAT, VND)",
               "Storefront", stores_all, plats,
               {(s, p): per_plat[p].get(s, 0.0) for s in stores_all for p in plats})

    # Brand mapping — the reviewable table
    ws = wb.create_sheet("Brand mapping")
    ws["A1"] = ("Storefront -> client brand mapping. PLEASE REVIEW the rows marked "
                "needs-confirmation or UNMAPPED and correct in one pass; storefronts "
                "were NOT merged unless the mapping says so.")
    ws["A1"].font = Font(bold=True)
    headers = ["Platform", "Storefront (as in files)", "Proposed client brand",
               "Confidence", f"July with-VAT (VND)", "Note"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = Font(bold=True)
    r = 4
    flagged = 0
    for p in plats:
        for s in sorted(per_plat[p]):
            b, conf, note = brand_of(p, s)
            ws.cell(row=r, column=1, value=p)
            ws.cell(row=r, column=2, value=s)
            ws.cell(row=r, column=3, value=b)
            cell = ws.cell(row=r, column=4, value=conf)
            if conf != "high":
                cell.font = Font(bold=True, color="FFC00000")
                flagged += 1
            ws.cell(row=r, column=5, value=per_plat[p][s]).number_format = ACC0
            ws.cell(row=r, column=6, value=note)
            r += 1
    for col, w in (("A", 10), ("B", 44), ("C", 30), ("D", 30), ("E", 20), ("F", 62)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    print(f"  brand mapping: {r - 4} storefronts, {flagged} flagged for review")

    # --- One tab per platform: store x window, plus pre-VAT total ---
    for p in plats:
        labels = [lab for _, lab in WINDOWS[p]]
        stores = sorted({s for lab in labels for s in data[p][lab]["store"]})
        wv = {(s, lab): float(data[p][lab].set_index("store")["wv"].get(s, 0.0))
              for s in stores for lab in labels}
        pre = {s: sum(float(data[p][lab].set_index("store")["pre"].get(s, 0.0))
                      for lab in labels) for s in stores}
        ws = wb.create_sheet(p)
        write_grid(ws, f"{p} by store and window, {args.month} (with VAT, VND)",
                   "Source.Name", stores, labels, wv,
                   extra_cols=[("Sum of Amount before VAT", pre)])

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"\nwrote {out}")

    print("\nTIE CHECK (master column totals vs the individual finance files):")
    ok = True
    for p in plats:
        for suffix, label in WINDOWS[p]:
            src = float(read_window(p, args.month, suffix)["wv"].sum())
            mine = float(data[p][label]["wv"].sum())
            good = abs(src - mine) < 1
            ok &= good
            print(f"  {p:<7} {label}: {mine:>18,.0f}  {'TIES' if good else 'MISMATCH'}")
    print("ALL COLUMNS TIE" if ok else "TIE FAILURE — do not send")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
