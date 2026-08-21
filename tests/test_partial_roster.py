"""`apply_partial_roster` names its stores (D3, 2026-08-21).

The all-or-nothing relaxation was register item D3's first clause: declaring a
window partial made EVERY expected store optional, so a genuinely forgotten
store was waved through with the legitimately absent ones. The function now
takes the declared list and relaxes only those names.

The contract these tests pin hardest is the LEGACY one: `optional_stores=None`
must union the whole expected list, byte-for-byte the old behaviour — it is what
`tools/devrun.py --partial-roster` and `tools/make_golden.py --partial-roster`
produce, and the four partial-roster goldens (2026-05_s1/s2/s3 shopee,
2026-05_w1 tiktok) are regenerated through exactly that path. If (a) fails, the
golden pipeline has drifted, not these tests.
"""

from __future__ import annotations

import pytest

pd = pytest.importorskip("pandas")

from src.errors import ReconHardStop          # noqa: E402
from src.ingest import check_stores           # noqa: E402
from src.pipeline import apply_partial_roster # noqa: E402
from src.runlog import RunLog


def _settings() -> dict:
    return {
        "expected_stores": {"tiktok": ["Abbott", "KAO", "Mars"]},
        "stores_optional": {"tiktok": ["Nutifood"]},
    }


def _frame(stores: list[str]) -> "pd.DataFrame":
    return pd.DataFrame({"store": stores, "amount": [1] * len(stores)})


def test_none_relaxes_the_whole_roster_the_legacy_and_golden_contract():
    settings = _settings()
    count = apply_partial_roster(settings, "tiktok", None)
    assert count == 3
    assert settings["stores_optional"]["tiktok"] == \
        ["Abbott", "KAO", "Mars", "Nutifood"], \
        "None = the pre-D3 blanket, and the config-level optional list survives"
    # Under the blanket, a window with ONE store passes — including one where a
    # store was genuinely forgotten, which is the defect this change closes.
    check_stores(_frame(["KAO"]), "orders", "tiktok", settings, RunLog())


def test_a_named_list_relaxes_only_those_stores():
    settings = _settings()
    count = apply_partial_roster(settings, "tiktok", ["Mars"])
    assert count == 1
    assert settings["stores_optional"]["tiktok"] == ["Mars", "Nutifood"]

    log = RunLog()
    # Mars declared absent: a window without it passes.
    check_stores(_frame(["Abbott", "KAO"]), "orders", "tiktok", settings, log)
    # KAO was NOT declared absent: its absence hard-stops again, naming it.
    with pytest.raises(ReconHardStop, match="KAO"):
        check_stores(_frame(["Abbott"]), "orders", "tiktok", settings, log)


def test_a_declared_store_the_roster_does_not_know_is_a_hard_stop():
    """A declaration written against a different roster (a repin, a rename) must
    fail HERE, naming the bad name — not one step later as a misleading
    'missing store' stop, and never as a silent skip."""
    settings = _settings()
    with pytest.raises(ReconHardStop, match="Pediasure"):
        apply_partial_roster(settings, "tiktok", ["Pediasure"])
    assert settings.get("stores_optional", {}).get("tiktok") in (None, ["Nutifood"]), \
        "a refused declaration must not have half-applied"


def test_the_unexpected_store_check_stays_armed_either_way():
    """The asymmetry that makes relaxation safe (D3/D23): a subset window is
    legitimate, a store nobody confirmed is not."""
    settings = _settings()
    apply_partial_roster(settings, "tiktok", None)
    with pytest.raises(ReconHardStop, match="Rogue"):
        check_stores(_frame(["KAO", "Rogue Store"]).assign(store=["KAO", "Rogue"]),
                     "orders", "tiktok", settings, RunLog())


def test_an_empty_list_relaxes_nothing():
    """[] is not None: it says 'partial, but I name no stores', and relaxing
    nothing is the only reading that cannot excuse a forgotten store."""
    settings = _settings()
    assert apply_partial_roster(settings, "tiktok", []) == 0
    with pytest.raises(ReconHardStop, match="Mars"):
        check_stores(_frame(["Abbott", "KAO"]), "orders", "tiktok", settings, RunLog())


# ---------------------------------------------------------------------------
# The workbook stamp (D3's third clause — the D46 deferral, closed)
# ---------------------------------------------------------------------------

def test_the_stamp_names_the_absent_stores():
    from src.finance_template import _roster_stamp

    stamp = _roster_stamp({"roster_relaxed": ["KAO", "Mars"]}, ["KAO"])
    assert stamp is not None
    assert "Mars" in stamp and "1 cửa hàng" in stamp, \
        "names, not only a count — the workbook must stand apart from the system"
    assert "KAO" not in stamp, "a relaxed store that IS present is not absent"


def test_the_stamp_is_relaxed_set_based_never_expected_minus_found():
    """A config-level stores_optional store absent from a NON-partial window must
    not stamp — the four non-partial lazada goldens are the proof this guards."""
    from src.finance_template import _roster_stamp

    assert _roster_stamp({"roster_relaxed": []}, ["KAO"]) is None
    assert _roster_stamp({}, []) is None


def test_a_declaration_the_window_outgrew_stamps_the_other_sentence():
    from src.finance_template import _roster_stamp

    stamp = _roster_stamp({"roster_relaxed": ["KAO"]}, ["KAO", "Mars"])
    assert stamp is not None and "đủ cửa hàng" in stamp, \
        "declared partial but nothing absent — the record needs re-checking"


def test_the_stamp_reaches_the_workbook_cells():
    """Through the real build_tiktok: the two cells exist iff the run relaxed
    something, and no `checks` entry is added either way (a checks entry would
    move fingerprint_digest alongside the workbook digest)."""
    from src.finance_template import ROSTER_STAMP_LABEL, build_tiktok
    from src.runlog import RunLog

    sku = pd.DataFrame({
        "store": ["KAO 8"] * 2,
        "order_id": ["A1", "A2"],
        "sku_id": ["S1", "S2"],
        "sku_name": ["N1", "N2"],
        "vat_factor": [1.08, 1.08],
        "unit_price_pre_vat": [100.0, 200.0],
        "quantity": [1, 1],
        "amount_pre_vat": [100.0, 200.0],
        "amount_with_vat": [108.0, 216.0],
        "actual_refund": [0.0, 0.0],
        "income_order_created_at": ["2026-05-01", "2026-05-02"],
        "statement_date": ["2026-05-03", "2026-05-04"],
        "check_status": ["", ""],
    })
    settings = {
        "vat_factors": {"rates": [1.05, 1.08, 1.10], "default": 1.08},
        "invoice_buckets": {"tiktok": {"match": {"merries": "Merries 8",
                                                 "kao": "KAO 8"},
                                       "default": "Others 8"}},
    }

    def pv_sum_cells(meta) -> dict:
        wb, checks = build_tiktok(sku.copy(), settings, dict(meta), RunLog())
        import io

        from openpyxl import load_workbook
        buf = io.BytesIO()
        wb.save(buf)
        ws = load_workbook(buf)["PV sum"]
        return {(c.coordinate): c.value for row in ws.iter_rows() for c in row
                if c.value is not None}, checks

    plain, plain_checks = pv_sum_cells({"label": ""})
    stamped, stamped_checks = pv_sum_cells(
        {"label": "", "roster_relaxed": ["KAO 8", "Mars"]})

    assert "A1" not in plain and "B1" not in plain
    assert stamped["A1"] == ROSTER_STAMP_LABEL
    assert "Mars" in stamped["B1"]
    assert {k: v for k, v in stamped.items() if k not in ("A1", "B1")} == plain, \
        "the stamp adds exactly two cells and moves nothing else"
    assert plain_checks == stamped_checks, \
        "the stamp must not add a checks entry — that would move fingerprint_digest"
