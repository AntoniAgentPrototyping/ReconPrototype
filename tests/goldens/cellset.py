"""Canonical workbook representation for the parity gate.

Deliberately depends on nothing but stdlib + openpyxl — no pandas, no polars.
The differ must not share a dataframe engine with either side it compares, or
an engine bug could cancel itself out.

Why not just hash the .xlsx bytes: openpyxl stamps `docProps/core.xml` with
created/modified timestamps and zip entry order is not guaranteed, so two
runs of identical data produce different files. Content is what matters.
"""

from __future__ import annotations

import datetime as _dt
import hashlib
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

CellValue = float | int | str | bool | None

# Type tags are derived from the Python value rather than openpyxl's
# data_type, because the bug worth catching is `"1000"` (text) sitting where
# `1000` (number) belongs — a defect number formats happily disguise.
TAG_BLANK = "z"
TAG_NUMBER = "n"
TAG_STRING = "s"
TAG_BOOL = "b"
TAG_DATE = "d"


def type_tag(value: CellValue) -> str:
    if value is None:
        return TAG_BLANK
    if isinstance(value, bool):  # before int — bool is an int subclass
        return TAG_BOOL
    if isinstance(value, (int, float)):
        return TAG_NUMBER
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        return TAG_DATE
    return TAG_STRING


@dataclass(frozen=True)
class Cell:
    tag: str
    value: CellValue
    number_format: str


@dataclass(frozen=True)
class Sheet:
    index: int
    name: str
    max_row: int
    max_col: int
    cells: dict[tuple[int, int], Cell]

    def ref(self, row: int, col: int) -> str:
        return f"{get_column_letter(col)}{row}"


@dataclass(frozen=True)
class CellSet:
    sheets: tuple[Sheet, ...]

    def by_name(self) -> dict[str, Sheet]:
        return {s.name: s for s in self.sheets}

    @property
    def names(self) -> tuple[str, ...]:
        return tuple(s.name for s in self.sheets)


def load_cellset(path: Path) -> CellSet:
    """Read a workbook into its canonical cellset.

    max_row/max_col are computed from the cells actually present, NOT from
    `ws.max_row`. Worksheet dimensions are an optional, and demonstrably
    unreliable, part of the format — a broken `<dimension>` tag in the June
    2026 TikTok exports is exactly why this project carries a calamine
    fallback (config/settings.yaml reader_engine). Deriving dimensions from
    content makes the gate independent of that.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = []
        for index, ws in enumerate(wb.worksheets):
            cells: dict[tuple[int, int], Cell] = {}
            max_row = max_col = 0
            for row in ws.iter_rows():
                for c in row:
                    if c.value is None:
                        # Blank cells are omitted, so a value-vs-blank
                        # mismatch surfaces as a key-set difference
                        # (NULLNESS) rather than being silently equal.
                        continue
                    key = (c.row, c.column)
                    cells[key] = Cell(
                        tag=type_tag(c.value),
                        value=c.value,
                        number_format=c.number_format or "General",
                    )
                    max_row = max(max_row, c.row)
                    max_col = max(max_col, c.column)
            sheets.append(Sheet(index, ws.title, max_row, max_col, cells))
        return CellSet(tuple(sheets))
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Serialization — floats via float.hex() so goldens round-trip bit-exactly.
# ---------------------------------------------------------------------------

def _encode(value: CellValue) -> object:
    if isinstance(value, float):
        return {"__f__": value.hex()}
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        return {"__dt__": value.isoformat()}
    return value


def _decode(value: object) -> CellValue:
    if isinstance(value, dict):
        if "__f__" in value:
            return float.fromhex(value["__f__"])
        if "__dt__" in value:
            return _dt.datetime.fromisoformat(value["__dt__"])
    return value  # type: ignore[return-value]


def dump_jsonl(cellset: CellSet, path: Path) -> None:
    """One JSON object per sheet. Line-oriented so a golden diff is readable
    in git and a single sheet can be streamed without parsing the whole file.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as fh:
        for s in cellset.sheets:
            payload = {
                "index": s.index,
                "name": s.name,
                "max_row": s.max_row,
                "max_col": s.max_col,
                "cells": [
                    [r, c, cell.tag, _encode(cell.value), cell.number_format]
                    for (r, c), cell in sorted(s.cells.items())
                ],
            }
            fh.write(json.dumps(payload, ensure_ascii=False, sort_keys=True) + "\n")


def load_jsonl(path: Path) -> CellSet:
    sheets = []
    with path.open("r", encoding="utf-8") as fh:
        for line in fh:
            if not line.strip():
                continue
            d = json.loads(line)
            cells = {
                (r, c): Cell(tag=tag, value=_decode(v), number_format=fmt)
                for r, c, tag, v, fmt in d["cells"]
            }
            sheets.append(Sheet(d["index"], d["name"], d["max_row"], d["max_col"], cells))
    return CellSet(tuple(sorted(sheets, key=lambda s: s.index)))


def as_cellset(source: Path | CellSet) -> CellSet:
    """Accept a live .xlsx, a stored golden .jsonl, or an in-memory CellSet."""
    if isinstance(source, CellSet):
        return source
    p = Path(source)
    return load_jsonl(p) if p.suffix == ".jsonl" else load_cellset(p)


# ---------------------------------------------------------------------------
# Digests — what gets committed. Hashes only: no values, no store names, no PII.
# ---------------------------------------------------------------------------

def _canonical_lines(sheet: Sheet) -> Iterator[str]:
    for (r, c), cell in sorted(sheet.cells.items()):
        value = cell.value.hex() if isinstance(cell.value, float) else repr(cell.value)
        yield f"{r}\t{c}\t{cell.tag}\t{value}\t{cell.number_format}"


def sheet_digest(sheet: Sheet) -> str:
    h = hashlib.sha256()
    h.update(f"{sheet.name}\t{sheet.max_row}\t{sheet.max_col}\n".encode())
    for line in _canonical_lines(sheet):
        h.update(line.encode() + b"\n")
    return h.hexdigest()


def manifest(cellset: CellSet) -> dict:
    """The committed artifact: one-way digests plus shape. Safe for git."""
    return {
        "schema": 1,
        "sheets": [
            {
                "index": s.index,
                "name": s.name,
                "max_row": s.max_row,
                "max_col": s.max_col,
                "cell_count": len(s.cells),
                "digest": sheet_digest(s),
            }
            for s in cellset.sheets
        ],
    }
