"""The month-end master's internal gate (M8 Phase 3, task 3.6).

`tools/build_master_summary.py` used to *print* "ALL COLUMNS TIE" and return an
exit code nobody read in CI. The claim is worth more as an assertion: every figure
in the master must be the same number the window it came from carries, because the
entire argument for this file is that it adds no arithmetic of its own.

Deliberately built on SYNTHETIC finance files. The real ones are client data and
live outside the repo, so a test that needed them could not run in CI — and the
properties under test (the forward-fill, the Lazada with-VAT recombination, the
coverage stamp, the grid totals) are structural, not data-dependent. The real-data
equivalent is `tools/build_master_summary.py --check`, which runs the same
`tie_rows` assertion over whatever is in `output/`.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

pytest.importorskip("pandas")
import pandas as pd                                                # noqa: E402
from openpyxl import Workbook                                      # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src import master_summary as ms                               # noqa: E402
from src.errors import ReconHardStop                               # noqa: E402

HEADER_ROW = ms.HEADER_ROW


def _finance_file(path: Path, tabs: dict[str, list[dict]], columns: list[str]) -> Path:
    """A workbook shaped like `finance_template`'s output: header on row 6."""
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in tabs.items():
        ws = wb.create_sheet(name)
        ws.cell(row=1, column=1, value="band row the template writes above the header")
        for j, col in enumerate(columns, start=1):
            ws.cell(row=HEADER_ROW, column=j, value=col)
        for i, row in enumerate(rows, start=HEADER_ROW + 1):
            for j, col in enumerate(columns, start=1):
                ws.cell(row=i, column=j, value=row.get(col))
    wb.save(path)
    return path


TT_COLS = ["Source.Name", "Amount before VAT", "Check total"]
LZ_COLS = ["Source.Name", "Amount"]


def test_a_repeat_sku_row_keeps_its_store(tmp_path):
    """The 'non repeat' convention blanks order-level columns on later SKU lines.

    Without the forward-fill every line after an order's first attributes to
    nothing, and the storefront's revenue silently shrinks — the quiet failure
    this whole file exists to prevent.
    """
    path = _finance_file(tmp_path / "f.xlsx", {"Xuat HD bt": [
        {"Source.Name": "Kao", "Amount before VAT": 100, "Check total": 108},
        {"Source.Name": None,  "Amount before VAT": 200, "Check total": 216},
        {"Source.Name": "Mars", "Amount before VAT": 50, "Check total": 54},
    ]}, TT_COLS)

    totals = ms.read_window(path, "TikTok")
    assert dict(zip(totals["store"], totals["pre"])) == {"kao": 300.0, "mars": 50.0}
    assert totals["wv"].sum() == pytest.approx(378.0)


def test_lazada_with_vat_is_recombined_from_the_tab_rate(tmp_path):
    """Lazada line tabs carry no with-VAT column; the tab NAME is the VAT factor."""
    path = tmp_path / "lz.xlsx"
    _finance_file(path, {
        "1.05": [{"Source.Name": "Curel", "Amount": 1000}],
        "1.08": [{"Source.Name": "Curel", "Amount": 2000}],
        "1.10": [{"Source.Name": "Kao", "Amount": 500}],
        "Fee buckets": [{"Source.Name": "not a line tab", "Amount": 999999}],
    }, LZ_COLS)

    totals = ms.read_window(path, "Lazada").set_index("store")
    assert totals.loc["curel", "pre"] == pytest.approx(3000.0)
    assert totals.loc["curel", "wv"] == pytest.approx(1000 * 1.05 + 2000 * 1.08)
    assert totals.loc["kao", "wv"] == pytest.approx(550.0)
    # 'Fee buckets' is not a rate-named tab and must contribute nothing.
    assert "not a line tab" not in totals.index


def test_a_file_with_no_line_tab_is_a_hard_stop(tmp_path):
    """Contributing a silent zero to a month-end total is the failure mode."""
    path = _finance_file(tmp_path / "empty.xlsx",
                         {"Summary": [{"Source.Name": "Kao", "Amount": 1}]}, LZ_COLS)
    with pytest.raises(ReconHardStop) as exc:
        ms.read_window(path, "TikTok")
    assert "Xuat HD bt" in str(exc.value)


# ---------------------------------------------------------------------------
# The gate itself
# ---------------------------------------------------------------------------

def _window(platform: str, period: str, stores: dict[str, tuple[float, float]]) -> ms.Window:
    totals = pd.DataFrame({"store": list(stores),
                           "pre": [v[0] for v in stores.values()],
                           "wv": [v[1] for v in stores.values()]})
    return ms.Window(platform=platform, period=period,
                     label=ms.window_label(period), totals=totals, source=period)


def _month() -> list[ms.Window]:
    return [
        _window("TikTok", "2026-07_w1", {"kao": (100.0, 108.0), "mars": (50.0, 54.0)}),
        _window("TikTok", "2026-07_w2", {"kao": (200.0, 216.0)}),
        # The sub-batch window the old hardcoded table omitted (register A5).
        _window("Shopee", "2026-07_s2", {"masan": (400.0, 432.0)}),
        _window("Shopee", "2026-07_s2x", {"masan": (25.0, 27.0)}),
        _window("Lazada", "2026-07_l1", {"curel": (10.0, 10.5)}),
    ]


def _grid_total(ws, header: str) -> float:
    """Sum a named column of a `_grid` sheet, from its data rows."""
    headers = {ws.cell(row=3, column=j).value: j
               for j in range(1, ws.max_column + 1)}
    col = headers[header]
    return sum(ws.cell(row=i, column=col).value or 0.0
               for i in range(4, ws.max_row))       # max_row is the =SUM() row


def test_every_master_figure_ties_to_the_window_it_came_from():
    """Task 3.6's internal gate, as an assertion rather than a printed line."""
    windows = _month()
    coverage = ms.Coverage(month="2026-07",
                           included=[(w.platform, w.period) for w in windows])
    wb = ms.build(coverage, windows)

    # 1. Every per-platform tab's with-VAT total equals its windows' sum.
    for platform in ("TikTok", "Shopee", "Lazada"):
        mine = sum(float(w.totals["wv"].sum()) for w in windows
                   if w.platform == platform)
        ws = wb[platform]
        got = sum(_grid_total(ws, w.label) for w in windows if w.platform == platform)
        assert got == pytest.approx(mine), f"{platform} tab does not tie to its windows"

    # 2. By-brand and by-storefront cover exactly the same money.
    grand = sum(float(w.totals["wv"].sum()) for w in windows)
    for sheet in ("By brand", "By storefront"):
        got = sum(_grid_total(wb[sheet], p) for p in ("TikTok", "Shopee", "Lazada"))
        assert got == pytest.approx(grand), f"{sheet} does not tie to the month"

    # 3. tie_rows agrees with the frames, per window.
    for row, w in zip(ms.tie_rows(windows), windows):
        assert row["wv"] == pytest.approx(float(w.totals["wv"].sum()))
        assert row["pre"] == pytest.approx(float(w.totals["pre"].sum()))
        assert row["stores"] == len(w.totals)


def test_a_sub_batch_window_is_not_dropped():
    """`s2x` is a real Shopee window. The hardcoded table it replaced had no row
    for it, and that table's own tie check re-read the same table."""
    windows = _month()
    wb = ms.build(ms.Coverage(month="2026-07"), windows)
    labels = {wb["Shopee"].cell(row=3, column=j).value
              for j in range(2, wb["Shopee"].max_column + 1)}
    assert {"s2", "s2x"} <= labels
    assert ms.window_label("2026-07_s2x") == "s2x"


# ---------------------------------------------------------------------------
# Coverage (task 3.5)
# ---------------------------------------------------------------------------

def test_a_partial_master_says_so_on_its_face():
    """A master that looks complete and is not is the failure this prevents."""
    windows = _month()
    coverage = ms.Coverage(
        month="2026-07",
        included=[(w.platform, w.period) for w in windows],
        missing=[("Lazada", "2026-07_l2", "the run stopped and produced no finance file")])
    wb = ms.build(coverage, windows)

    assert not coverage.complete
    assert wb.sheetnames[0] == "Coverage", "coverage must be the first thing seen"
    text = "\n".join(str(c.value) for row in wb["Coverage"].iter_rows() for c in row
                     if c.value is not None)
    assert "PARTIAL" in text
    assert "2026-07_l2" in text and "MISSING" in text
    # The banner is repeated on Summary: that is the tab people actually read.
    assert "PARTIAL" in str(wb["Summary"]["A2"].value)


def test_a_complete_master_does_not_cry_wolf():
    windows = _month()
    coverage = ms.Coverage(month="2026-07",
                           included=[(w.platform, w.period) for w in windows])
    wb = ms.build(coverage, windows)
    assert coverage.complete
    assert "PARTIAL" not in coverage.headline()
    assert wb["Summary"]["A2"].value is None


def test_an_unmapped_storefront_is_flagged_and_never_merged():
    """Merging two storefronts that are separate client entities invoices one
    client for another's revenue, so nothing merges without an explicit row."""
    windows = _month()
    brand_map = ms.parse_brand_map(
        "platform,storefront,client_brand,confidence,note\n"
        "tiktok,kao,KAO,high,\n")
    wb = ms.build(ms.Coverage(month="2026-07"), windows, brand_map)

    rows = {(wb["Brand mapping"].cell(row=i, column=2).value,
             wb["Brand mapping"].cell(row=i, column=3).value,
             wb["Brand mapping"].cell(row=i, column=4).value)
            for i in range(4, wb["Brand mapping"].max_row + 1)}
    assert ("kao", "KAO", "high") in rows
    assert ("mars", "mars", "UNMAPPED (kept as storefront)") in rows
